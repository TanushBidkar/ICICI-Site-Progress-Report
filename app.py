from flask import Flask, render_template, request, jsonify, send_file
import firebase_admin
from firebase_admin import credentials, storage
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, Alignment
from werkzeug.middleware.proxy_fix import ProxyFix
from openpyxl.styles import Border, Side
from PIL import Image
import io
import os
from datetime import datetime
import json
import base64
import uuid
from PIL import Image as PILImage
from openpyxl.styles import PatternFill  # Add this line
import os
from dotenv import load_dotenv
import gc  # Add this line at the top with other imports
import bcrypt
from datetime import datetime, timezone, timedelta

app = Flask(__name__)
app.wsgi_app = ProxyFix(app.wsgi_app)

app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024  # 50MB limit
load_dotenv()

# Firebase Configuration
firebase_config = {
    "apiKey": os.getenv('FIREBASE_API_KEY'),
    "authDomain": os.getenv('FIREBASE_AUTH_DOMAIN'),
    "projectId": os.getenv('FIREBASE_PROJECT_ID'),
    "storageBucket": os.getenv('FIREBASE_STORAGE_BUCKET'),
    "messagingSenderId": os.getenv('FIREBASE_MESSAGING_SENDER_ID'),
    "appId": os.getenv('FIREBASE_APP_ID')
}

# Initialize Firebase
if not firebase_admin._apps:
    cred = credentials.Certificate(os.getenv('SERVICE_ACCOUNT_KEY_PATH', 'serviceAccountKey.json'))
    firebase_admin.initialize_app(cred, {
        'storageBucket': firebase_config['storageBucket']
    })

bucket = storage.bucket()


# ✅ ADD THIS LINE:
from firebase_admin import firestore
db = firestore.client()
# Region-wise reviewers configuration
REGION_REVIEWERS = {
    'West': [
        {'email': 'richarddsilva@company.com', 'name': 'Richard Dsilva'},
        {'email': 'abhaypatwa@company.com', 'name': 'Abhay Patwa'},
        {'email': 'niwantghadge@company.com', 'name': 'Niwant Ghadge'},
        {'email': 'chaman.gotya@cushwake.com', 'name': 'Chaman Gotya'}
    ],
    'East': [
        {'email': 'subhomoybhakat@company.com', 'name': 'Subhomoy Bhakat'},
        {'email': 'dipayansutar@company.com', 'name': 'Dipayan Sutar'},
        {'email': 'souryabratachatterjee@company.com', 'name': 'Sourya Chatterjee'}
    ],
    'South': [
        {'email': 'adhithyaram@company.com', 'name': 'Adithyaraman Natarajan'},
        {'email': 'arjunkumarm@company.com', 'name': 'Arjun Kumar M'}
    ],
    'North': [
        {'email': 'marufkhan@company.com', 'name': 'Maruf Khan'},
        {'email': 'amandeepmaurya@company.com', 'name': 'Amandeep Mourya'}
    ]
}

# Template path
TEMPLATE_PATH = 'templates/excel_template.xlsx'

def calculate_row_height(text, font_size=12, cell_width=50):
    """Calculate appropriate row height based on text length"""
    if not text:
        return 15  # Default height for empty cells
    
    # Approximate characters per line based on cell width
    chars_per_line = cell_width * 1.5  # Rough estimate
    
    # Calculate number of lines needed
    lines = len(text) / chars_per_line
    if lines < 1:
        lines = 1
    
    # Calculate height (font_size * 1.5 per line + padding)
    height = (font_size * 1.5 * lines) + 5
    
    # Minimum height
    if height < 15:
        height = 15
    
    # Maximum height (to avoid extremely tall rows)
    if height > 150:
        height = 150
    
    return height

def unmerge_and_write(ws, cell_address, value):
    """Safely write to a cell, handling merged cells"""
    try:
        # Check if cell is part of a merged range
        cell = ws[cell_address]
        for merged_range in list(ws.merged_cells.ranges):  # Convert to list to avoid modification during iteration
            if cell.coordinate in merged_range:
                # Unmerge the range
                ws.unmerge_cells(str(merged_range))
                # Write to the top-left cell of the merged range
                top_left_cell = merged_range.start_cell
                top_left_cell.value = value
                # Re-merge the cells
                ws.merge_cells(str(merged_range))
                return
        # If not merged, write directly
        cell.value = value
    except Exception as e:
        print(f"Error writing to {cell_address}: {e}")
        try:
            ws[cell_address].value = value
        except:
            pass
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/check-existing', methods=['POST'])
def check_existing():
    """Check if SOL ID and Visit No already exists"""
    data = request.json
    sol_id = data.get('sol_id')
    visit_no = data.get('visit_no')
    
    # Check in Firebase Storage
    blob_path = f"ICICI_Site_Progress_Report/{sol_id}/Visit_{visit_no}/data.json"
    blob = bucket.blob(blob_path)
    
    if blob.exists():
        # Download existing data
        existing_data = json.loads(blob.download_as_string())
        return jsonify({
            'exists': True,
            'data': existing_data  # Return ALL form data
        })
    
    return jsonify({'exists': False})
@app.route('/get-existing-images', methods=['POST'])
def get_existing_images():
    """Get existing images for a visit"""
    data = request.json
    sol_id = data.get('sol_id')
    visit_no = data.get('visit_no')
    photo_count = int(data.get('photo_count', 0))
    
    images = {}
    
    for i in range(photo_count):
        try:
            blob_path = f"ICICI_Site_Progress_Report/{sol_id}/Visit_{visit_no}/photo_{i}.jpg"
            blob = bucket.blob(blob_path)
            
            if blob.exists():
                # Get image as base64
                image_bytes = blob.download_as_bytes()
                base64_image = base64.b64encode(image_bytes).decode('utf-8')
                images[f'photo_image_{i}'] = base64_image
        except Exception as e:
            print(f"Error loading image {i}: {e}")
    
    return jsonify({'images': images})

@app.route('/get-project-name', methods=['POST'])
def get_project_name():
    """Get project name based on SOL ID"""
    data = request.json
    sol_id = data.get('sol_id')
    
    # Search for any visit with this SOL ID
    prefix = f"ICICI_Site_Progress_Report/{sol_id}/"
    blobs = bucket.list_blobs(prefix=prefix, max_results=1)
    
    for blob in blobs:
        if 'data.json' in blob.name:
            existing_data = json.loads(blob.download_as_string())
            return jsonify({
                'found': True,
                'project_name': existing_data.get('project_name', '')
            })
    
    return jsonify({'found': False})

@app.route('/get-image-from-firebase', methods=['POST'])
def get_image_from_firebase():
    """Fetch a single image from Firebase as base64"""
    try:
        data = request.json
        sol_id = data.get('sol_id')
        visit_no = data.get('visit_no')
        image_type = data.get('image_type')  # 'work', 'qual', 'make'
        image_index = data.get('image_index')
        
        blob_path = f"ICICI_Site_Progress_Report/{sol_id}/Visit_{visit_no}/{image_type}_{image_index}.jpg"
        blob = bucket.blob(blob_path)
        
        if blob.exists():
            image_bytes = blob.download_as_bytes()
            base64_image = base64.b64encode(image_bytes).decode('utf-8')
            return jsonify({'image_base64': base64_image})
        else:
            return jsonify({'image_base64': None}), 404
            
    except Exception as e:
        print(f"Error fetching image: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/submit-report', methods=['POST'])
def submit_report():
    """Process and save the complete report"""
    try:
        # Get form data
        form_data = request.form.to_dict()
        files = request.files
        
        sol_id = form_data['sol_id']
        visit_no = form_data['visit_no']
        project_name = form_data['project_name']
        
        # Check if this is an update (existing visit)
        blob_path = f"ICICI_Site_Progress_Report/{sol_id}/Visit_{visit_no}/data.json"
        existing_blob = bucket.blob(blob_path)
        is_update = existing_blob.exists()
        
        # Load template
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        
        # Fill Progress Report Sheet
        fill_progress_report(wb, form_data)
        
        # Save images to Firebase (before filling photographs sheet)
        save_images_to_firebase(sol_id, visit_no, form_data, files)
        
        # Fill Site Visit Photographs Sheet
        fill_photographs_sheet(wb, form_data, files, sol_id, visit_no)
        
        # Fill Quality and Critical Challenges Sheet
        fill_quality_sheet(wb, form_data)
        
        # Save to memory
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
       # Upload to Firebase with metadata to enable download in console
        report_blob_path = f"ICICI_Site_Progress_Report/{sol_id}/Visit_{visit_no}/report.xlsx"
        report_blob = bucket.blob(report_blob_path)
        
        # ✅ Set metadata with content disposition to enable download
        report_blob.metadata = {
            'firebaseStorageDownloadTokens': str(uuid.uuid4())  # Generate unique token
        }
        report_blob.content_disposition = 'attachment; filename="report.xlsx"'
        
        report_blob.upload_from_file(
            excel_buffer, 
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # ✅ ADD THIS: Make the file publicly accessible
        report_blob.make_public()
        
       
        
        # Save JSON data (OVERWRITE existing)
        json_blob = bucket.blob(blob_path)
        json_blob.upload_from_string(json.dumps(form_data), content_type='application/json')
        
        action = "updated" if is_update else "created"
        print(f"Report {action} successfully for SOL ID: {sol_id}, Visit: {visit_no}")
        
        # Return download
        excel_buffer.seek(0)
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=f"{project_name}_Visit_{visit_no}_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"Submit report error: {str(e)}")
        return jsonify({'error': str(e)}), 500

def save_images_to_firebase(sol_id, visit_no, form_data, files):
    """Save uploaded images to Firebase Storage - MEMORY OPTIMIZED"""
    
    def process_and_upload_image(file_key, image_file, blob_path):
        """Helper to process single image and clean up memory"""
        try:
            img = Image.open(image_file)
            
            # ✅ Aggressive resize to reduce memory
            max_size = (800, 600)
            img.thumbnail(max_size, Image.Resampling.LANCZOS)
            
            img_buffer = io.BytesIO()
            img.save(img_buffer, format='JPEG', quality=75, optimize=True)  # Lower quality
            img_buffer.seek(0)
            
            blob = bucket.blob(blob_path)
            blob.upload_from_file(img_buffer, content_type='image/jpeg')
            
            # ✅ Immediate cleanup
            img.close()
            img_buffer.close()
            del img
            del img_buffer
            gc.collect()  # Force garbage collection
            
            return True
        except Exception as e:
            print(f"Error saving {file_key}: {e}")
            return False
    
    # Work Progress Images
    work_count = int(form_data.get('work_count', 0))
    for i in range(work_count):
        file_key = f'work_image_{i}'
        if file_key in files:
            blob_path = f"ICICI_Site_Progress_Report/{sol_id}/Visit_{visit_no}/work_{i}.jpg"
            process_and_upload_image(file_key, files[file_key], blob_path)
    
    # Quality Images
    quality_count = int(form_data.get('quality_count', 0))
    for i in range(quality_count):
        file_key = f'qual_image_{i}'
        if file_key in files:
            blob_path = f"ICICI_Site_Progress_Report/{sol_id}/Visit_{visit_no}/qual_{i}.jpg"
            process_and_upload_image(file_key, files[file_key], blob_path)
    
    # Make/Model Images
    make_count = int(form_data.get('make_count', 0))
    for i in range(make_count):
        file_key = f'make_image_{i}'
        if file_key in files:
            blob_path = f"ICICI_Site_Progress_Report/{sol_id}/Visit_{visit_no}/make_{i}.jpg"
            process_and_upload_image(file_key, files[file_key], blob_path)
def fill_progress_report(wb, data):
    """Fill the Progress Report sheet"""
    ws = wb['Progress Report']
    
    # Fill header information
    unmerge_and_write(ws, 'B2', data.get('project_name', ''))  # Project name
    
    # Branch Area should show "Branch Area: X" format
    branch_area = data.get('branch_area', '')
    unmerge_and_write(ws, 'G2', f'Branch Area: {branch_area}')   # Branch Area with label
    
    unmerge_and_write(ws, 'B3', data.get('branch_code', ''))   # Branch Code
    unmerge_and_write(ws, 'E3', data.get('date_of_visit', '')) # Date of visit (goes in E3)
    
    # Visit no should show "Visit No: X" format
    visit_no = data.get('visit_no', '')
    unmerge_and_write(ws, 'G3', f'Visit No: {visit_no}')      # Visit no with label
    
    # Get current visit number to know which column to fill
    current_visit = int(data.get('visit_no', '1'))
    
    # Fill Non-Internal Work + Civil items (rows 9-19: all items from a to k)
    # Fill Non-Internal Work + Civil items (rows 9-19: all items from a to k)
    non_internal_and_civil_items = [
        'demolition', 'block_work', 'internal_plaster', 'rcc_wall', 'pcc_work', 'pop_punning',
        'waterproofing', 'flooring', 'dado', 'painting', 'plumbing'
    ]

    for idx, item in enumerate(non_internal_and_civil_items, start=9):
        for visit in range(1, 5):
            cell_key = f"{item}_visit{visit}"
            if cell_key in data and data[cell_key]:
                col = chr(ord('B') + visit)  # Visit 1 = C, Visit 2 = D, etc.
                value = data[cell_key]
                if value and str(value).strip():
                    cell = ws[f'{col}{idx}']
                    cell.value = float(value) / 100
                    cell.number_format = '0%'
                    cell.alignment = Alignment(horizontal='right', vertical='center')  # ✅ RIGHT align

        remark_key = f"{item}_remark"
        if remark_key in data:
            unmerge_and_write(ws, f'G{idx}', data[remark_key])
    
    
    
    # Fill Carpentry Work
    carpentry_items = ['partition', 'paneling', 'door_window', 'false_ceiling', 
                       'loose_furniture', 'alum_skirting', 'window_blind', 'signage']
    for idx, item in enumerate(carpentry_items, start=22):
        for visit in range(1, 5):
            cell_key = f"{item}_visit{visit}"
            if cell_key in data and data[cell_key]:
                col = chr(ord('B') + visit)
                value = data[cell_key]
                if value and str(value).strip():
                    cell = ws[f'{col}{idx}']
                    cell.value = float(value) / 100
                    cell.number_format = '0%'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
        
        remark_key = f"{item}_remark"
        if remark_key in data:
            unmerge_and_write(ws, f'G{idx}', data[remark_key])  # Remarks in column G
    
    # Fill Electrical Work
    electrical_items = ['pipe_conduit', 'raceway', 'wiring', 'fixtures', 'main_lt']
    for idx, item in enumerate(electrical_items, start=32):
        for visit in range(1, 5):
            cell_key = f"{item}_visit{visit}"
            if cell_key in data and data[cell_key]:
                col = chr(ord('B') + visit)
                value = data[cell_key]
                if value and str(value).strip():
                    cell = ws[f'{col}{idx}']
                    cell.value = float(value) / 100
                    cell.number_format = '0%'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
        
        remark_key = f"{item}_remark"
        if remark_key in data:
            unmerge_and_write(ws, f'G{idx}', data[remark_key])  # Remarks in column G

    # Fill HVAC Work
    hvac_items = ['hvac_indoor', 'hvac_wiring', 'hvac_outdoor']
    for idx, item in enumerate(hvac_items, start=39):  # Adjust row number based on your Excel template
        for visit in range(1, 5):
            cell_key = f"{item}_visit{visit}"
            if cell_key in data and data[cell_key]:
                col = chr(ord('B') + visit)
                value = data[cell_key]
                if value and str(value).strip():
                    cell = ws[f'{col}{idx}']
                    cell.value = float(value) / 100
                    cell.number_format = '0%'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
        
        remark_key = f"{item}_remark"
        if remark_key in data:
            unmerge_and_write(ws, f'G{idx}', data[remark_key])
    
    # Fill CMS Work
    cms_items = ['cms_pipe', 'cms_wiring', 'cms_fixture']
    for idx, item in enumerate(cms_items, start=44):  # Adjust row number based on your Excel template
        for visit in range(1, 5):
            cell_key = f"{item}_visit{visit}"
            if cell_key in data and data[cell_key]:
                col = chr(ord('B') + visit)
                value = data[cell_key]
                if value and str(value).strip():
                    cell = ws[f'{col}{idx}']
                    cell.value = float(value) / 100
                    cell.number_format = '0%'
                    cell.alignment = Alignment(horizontal='right', vertical='center')
        
        remark_key = f"{item}_remark"
        if remark_key in data:
            unmerge_and_write(ws, f'G{idx}', data[remark_key])

        
    current_row = 48  # Start at row 48 (after one blank row)

# Get the actual count of "Others" items from form data
    other_count = int(data.get('other_count', 0))

# Define border style FIRST (before using it)
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )

    if other_count > 0:  # Only add "Others" section if user has items
    # Insert section header row at 41
        ws.insert_rows(current_row, 1)
    
    # Add "V" in column A
        ws[f'A{current_row}'] = 'VI'
        ws[f'A{current_row}'].font = Font(name='Calibri', size=11, bold=True)
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')

        # Add "Others" in column B ONLY (like "Electrical" format)
        ws[f'B{current_row}'] = 'Others'
        ws[f'B{current_row}'].font = Font(name='Calibri', size=11, bold=True)
        ws[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center')

        # Add borders to row 41
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            ws[f'{col}{current_row}'].border = thin_border
    
        current_row += 1  # Move to next row for first item (row 42)
    
    # Now add user's "Other" items with a, b, c... numbering
        actual_item_count = 0
        for i in range(other_count):
            item_name = data.get(f'other_item_{i}', '')
            if item_name and item_name.strip():
            # Insert a new row
                ws.insert_rows(current_row, 1)

                # Add serial number (a, b, c, d...)
                ws[f'A{current_row}'] = chr(97 + actual_item_count)
                ws[f'A{current_row}'].font = Font(name='Calibri', size=11)
                ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')

                # Write item name in column B (NOT BOLD, NORMAL)
                ws[f'B{current_row}'] = item_name
                ws[f'B{current_row}'].font = Font(name='Calibri', size=11)
                ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center')

                # Fill percentages for ALL visits (1 to 4)
                for visit in range(1, 5):
                    cell_key = f"other_item_{i}_visit{visit}"
                    if cell_key in data and data[cell_key]:
                        col = chr(ord('B') + visit)  # Visit 1=C, 2=D, 3=E, 4=F
                        value = data[cell_key]
                        if value and str(value).strip():
                            cell = ws[f'{col}{current_row}']
                            cell.value = float(value) / 100
                            cell.number_format = '0%'
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                    
                    
            
            # Fill remark in column G (NOT BOLD, NORMAL)
                remark_key = f"other_item_{i}_remark"
                if remark_key in data:
                    ws[f'G{current_row}'] = data[remark_key]
                    ws[f'G{current_row}'].font = Font(name='Calibri', size=11)
                    ws[f'G{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
            
            # Add borders to this row
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                    ws[f'{col}{current_row}'].border = thin_border
            
                actual_item_count += 1
                current_row += 1

# Leave ONE BLANK ROW after Others section
    current_row += 1

# NOW insert "Prepared By" and "Checked By" row
    ws.insert_rows(current_row, 1)

    # Merge E and F for "Prepared By:"
    ws.merge_cells(f'E{current_row}:F{current_row}')
    ws[f'E{current_row}'] = 'Prepared By:'
    ws[f'E{current_row}'].font = Font(name='Calibri', size=11, bold=True)
    ws[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center')

    # "Checked By:" in column G
    ws[f'G{current_row}'] = 'Checked By:'
    ws[f'G{current_row}'].font = Font(name='Calibri', size=11, bold=True)
    ws[f'G{current_row}'].alignment = Alignment(horizontal='center', vertical='center')

    # Add borders to Prepared By row
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws[f'{col}{current_row}'].border = thin_border

    # Add one more row for actual names (Ram and Tanush)
    current_row += 1
    ws.insert_rows(current_row, 1)

    # Merge E and F for prepared by name
    ws.merge_cells(f'E{current_row}:F{current_row}')
    ws[f'E{current_row}'] = data.get('prepared_by', '')
    ws[f'E{current_row}'].font = Font(name='Calibri', size=11, bold=True)
    ws[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center')

    # Checked by name in column G
    ws[f'G{current_row}'] = data.get('checked_by', '')
    ws[f'G{current_row}'].font = Font(name='Calibri', size=11, bold=True)
    ws[f'G{current_row}'].alignment = Alignment(horizontal='center', vertical='center')

    # Add borders to names row (FINAL ROW OF TABLE)
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
        ws[f'{col}{current_row}'].border = thin_border

    # ✅✅✅ NEW CODE: Remove all borders and formatting from rows below ✅✅✅
    # Clear any existing formatting from rows after the table ends
    max_clear_row = current_row + 10000  # Clear next 20 rows to be safe

    for clear_row in range(current_row + 1, max_clear_row + 1):
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            cell = ws[f'{col}{clear_row}']
            cell.border = Border()  # Remove all borders
            cell.value = None  # Clear any value
            cell.font = Font(name='Calibri', size=11)  # Reset to default font
            cell.alignment = Alignment(horizontal='general', vertical='bottom')  # Default alignment

    # Unmerge any merged cells below the table
    for merged_range in list(ws.merged_cells.ranges):
        # If merged range starts after our table, unmerge it
        if merged_range.min_row > current_row:
            ws.unmerge_cells(str(merged_range))



def fill_photographs_sheet(wb, data, files, sol_id, visit_no):
    """Fill Site visit Photographs - DYNAMIC INSERTION (NO HARDCODED ROWS)"""
    ws = wb['Site visit Photographs']
    
    # Define border style
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # ✅ SET EXACT COLUMN WIDTHS (per user specifications)
    ws.column_dimensions['A'].width = 31.89  # Sr. No. (294 px)
    ws.column_dimensions['B'].width = 50.22  # Critical Milestones (459 px)
    ws.column_dimensions['C'].width = 50.22  # Description (459 px)
    ws.column_dimensions['D'].width = 60  # Photographs (540 px)
    ws.column_dimensions['E'].width = 46.89  # Remarks (429 px)
    
    # ✅ DELETE ALL ROWS FROM ROW 9 ONWARDS (clear template placeholders)
    max_row = ws.max_row
    if max_row >= 9:
        ws.delete_rows(9, max_row - 8)
    
    current_row = 9  # Start fresh after "Work Progress" header (row 8)
    
    # ====================
    # 1. WORK PROGRESS SECTION
    # ====================
    work_count = int(data.get('work_count', 0))
    
    if work_count > 0:
        for i in range(work_count):
            ws.insert_rows(current_row, 1)
            ws.row_dimensions[current_row].height = 249.60
            
            # Add borders
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{col}{current_row}'].border = thin_border
            
            # Serial Number
            ws[f'A{current_row}'].value = i + 1
            ws[f'A{current_row}'].font = Font(name='Calibri', size=20, bold=True)
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Milestone
            ws[f'B{current_row}'].value = data.get(f'work_milestone_{i}', '')
            ws[f'B{current_row}'].font = Font(name='Calibri', size=24)
            ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Description
            ws[f'C{current_row}'].value = data.get(f'work_desc_{i}', '')
            ws[f'C{current_row}'].font = Font(name='Calibri', size=24)
            ws[f'C{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # Remarks
            ws[f'E{current_row}'].value = data.get(f'work_remark_{i}', '')
            ws[f'E{current_row}'].font = Font(name='Calibri', size=24)
            ws[f'E{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # IMAGE
            file_key = f'work_image_{i}'
            img_stream = None
            
            if file_key in files and files[file_key].filename != '':
                try:
                    files[file_key].seek(0)
                    img_stream = files[file_key]
                except:
                    pass
            
            if img_stream is None:
                try:
                    blob_path = f"ICICI_Site_Progress_Report/{sol_id}/Visit_{visit_no}/work_{i}.jpg"
                    blob = bucket.blob(blob_path)
                    if blob.exists():
                        img_bytes = blob.download_as_bytes()
                        img_stream = io.BytesIO(img_bytes)
                except Exception as e:
                    print(f"Error fetching work image {i}: {e}")
            
            if img_stream:
                try:
                    pil_img = PILImage.open(img_stream)
                    pil_img = pil_img.resize((430, 320), PILImage.Resampling.LANCZOS)
                    
                    img_byte_arr = io.BytesIO()
                    pil_img.save(img_byte_arr, format='PNG')
                    img_byte_arr.seek(0)
                    
                    xl_img = XLImage(img_byte_arr)
                    xl_img.anchor = f'D{current_row}'
                    ws.add_image(xl_img)
                    
                    ws[f'D{current_row}'].value = ""
                    
                except Exception as e:
                    print(f"Error inserting work image {i}: {e}")
                    ws[f'D{current_row}'].value = "Error loading image"
                    ws[f'D{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            else:
                ws[f'D{current_row}'].value = "[No Image]"
                ws[f'D{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            current_row += 1
    
    # ====================
    # 2. QUALITY OBSERVATION SECTION
    # ====================
    quality_count = int(data.get('quality_count', 0))
    
    if quality_count > 0:
        # Insert section title row
        ws.insert_rows(current_row, 1)
        ws.merge_cells(f'B{current_row}:E{current_row}')
        ws[f'B{current_row}'] = 'Quality Observation'
        ws.row_dimensions[current_row].height = 28.20  # Title row: 49 pixels
        ws[f'B{current_row}'].font = Font(name='Calibri', size=20, bold=True)
        ws[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{current_row}'].border = thin_border
        
        current_row += 1
        
        # Insert header row
        ws.insert_rows(current_row, 1)
        ws[f'A{current_row}'] = 'Sr. No.'
        ws[f'B{current_row}'] = 'Items'
        ws[f'C{current_row}'] = 'Description'
        ws[f'D{current_row}'] = 'Photographs'
        ws[f'E{current_row}'] = 'Remarks'
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{current_row}'].font = Font(name='Calibri', size=20, bold=True)
            ws[f'{col}{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'{col}{current_row}'].border = thin_border
        
        current_row += 1
        
        # Insert quality entries
        for i in range(quality_count):
            ws.insert_rows(current_row, 1)
            ws.row_dimensions[current_row].height = 249.60
            
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{col}{current_row}'].border = thin_border
            
            ws[f'A{current_row}'].value = i + 1
            ws[f'A{current_row}'].font = Font(name='Calibri', size=20, bold=True)
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws[f'B{current_row}'].value = data.get(f'qual_item_{i}', '')
            ws[f'B{current_row}'].font = Font(name='Calibri', size=24)
            ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            ws[f'C{current_row}'].value = data.get(f'qual_desc_{i}', '')
            ws[f'C{current_row}'].font = Font(name='Calibri', size=24)
            ws[f'C{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            ws[f'E{current_row}'].value = data.get(f'qual_remark_{i}', '')
            ws[f'E{current_row}'].font = Font(name='Calibri', size=24)
            ws[f'E{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # IMAGE
            file_key = f'qual_image_{i}'
            img_stream = None
            
            if file_key in files and files[file_key].filename != '':
                try:
                    files[file_key].seek(0)
                    img_stream = files[file_key]
                except:
                    pass
            
            if img_stream is None:
                try:
                    blob_path = f"ICICI_Site_Progress_Report/{sol_id}/Visit_{visit_no}/qual_{i}.jpg"
                    blob = bucket.blob(blob_path)
                    if blob.exists():
                        img_bytes = blob.download_as_bytes()
                        img_stream = io.BytesIO(img_bytes)
                except Exception as e:
                    print(f"Error fetching qual image {i}: {e}")
            
            if img_stream:
                try:
                    pil_img = PILImage.open(img_stream)
                    pil_img = pil_img.resize((430, 320), PILImage.Resampling.LANCZOS)
                    
                    img_byte_arr = io.BytesIO()
                    pil_img.save(img_byte_arr, format='PNG')
                    img_byte_arr.seek(0)
                    
                    xl_img = XLImage(img_byte_arr)
                    xl_img.anchor = f'D{current_row}'
                    ws.add_image(xl_img)
                    
                    ws[f'D{current_row}'].value = ""
                    
                except Exception as e:
                    print(f"Error inserting qual image {i}: {e}")
                    ws[f'D{current_row}'].value = "Error loading image"
                    ws[f'D{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            else:
                ws[f'D{current_row}'].value = "[No Image]"
                ws[f'D{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            current_row += 1
    
    # ====================
    # 3. MAKE/MODEL INSPECTION SECTION
    # ====================
    make_count = int(data.get('make_count', 0))
    
    if make_count > 0:
        # Insert section title
        ws.insert_rows(current_row, 1)
        ws.merge_cells(f'B{current_row}:E{current_row}')
        ws[f'B{current_row}'] = 'Make / Model Inspection'
        ws.row_dimensions[current_row].height = 28.20  # Title row: 49 pixels
        ws[f'B{current_row}'].font = Font(name='Calibri', size=20, bold=True)
        ws[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{current_row}'].border = thin_border
        
        current_row += 1
        
        # Insert header row
        ws.insert_rows(current_row, 1)
        ws[f'A{current_row}'] = 'Sr. No.'
        ws[f'B{current_row}'] = 'Items'
        ws[f'C{current_row}'] = 'Make/Model Observed'
        ws[f'D{current_row}'] = 'Photographs'
        ws[f'E{current_row}'] = 'Remarks'
        
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{current_row}'].font = Font(name='Calibri', size=20, bold=True)
            ws[f'{col}{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'{col}{current_row}'].border = thin_border
        
        current_row += 1
        
        # Insert make/model entries
        for i in range(make_count):
            ws.insert_rows(current_row, 1)
            ws.row_dimensions[current_row].height = 249.60
            
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws[f'{col}{current_row}'].border = thin_border
            
            ws[f'A{current_row}'].value = i + 1
            ws[f'A{current_row}'].font = Font(name='Calibri', size=20, bold=True)
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            ws[f'B{current_row}'].value = data.get(f'make_item_{i}', '')
            ws[f'B{current_row}'].font = Font(name='Calibri', size=24)
            ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            ws[f'C{current_row}'].value = data.get(f'make_observed_{i}', '')
            ws[f'C{current_row}'].font = Font(name='Calibri', size=24)
            ws[f'C{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            ws[f'E{current_row}'].value = data.get(f'make_remark_{i}', '')
            ws[f'E{current_row}'].font = Font(name='Calibri', size=24)
            ws[f'E{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            
            # IMAGE
            file_key = f'make_image_{i}'
            img_stream = None
            
            if file_key in files and files[file_key].filename != '':
                try:
                    files[file_key].seek(0)
                    img_stream = files[file_key]
                except:
                    pass
            
            if img_stream is None:
                try:
                    blob_path = f"ICICI_Site_Progress_Report/{sol_id}/Visit_{visit_no}/make_{i}.jpg"
                    blob = bucket.blob(blob_path)
                    if blob.exists():
                        img_bytes = blob.download_as_bytes()
                        img_stream = io.BytesIO(img_bytes)
                except Exception as e:
                    print(f"Error fetching make image {i}: {e}")
            
            if img_stream:
                try:
                    pil_img = PILImage.open(img_stream)
                    pil_img = pil_img.resize((430, 320), PILImage.Resampling.LANCZOS)
                    
                    img_byte_arr = io.BytesIO()
                    pil_img.save(img_byte_arr, format='PNG')
                    img_byte_arr.seek(0)
                    
                    xl_img = XLImage(img_byte_arr)
                    xl_img.anchor = f'D{current_row}'
                    ws.add_image(xl_img)
                    
                    ws[f'D{current_row}'].value = ""
                    
                except Exception as e:
                    print(f"Error inserting make image {i}: {e}")
                    ws[f'D{current_row}'].value = "Error loading image"
                    ws[f'D{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            else:
                ws[f'D{current_row}'].value = "[No Image]"
                ws[f'D{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            current_row += 1


def fill_quality_sheet(wb, data):
    """Fill the Quality and Critical Challenges sheet - REMOVES EMPTY ROWS"""
    ws = wb['Quality and Critical challenges']
    
    # Define border style
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # Define Green color (Accent 6, Lighter 80%) - RGB: E2EFDA
    green_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    
    # Fill header information (rows 2-9) with Calibri 12
    ws['B2'] = data.get('quality_project_title', '')
    ws['B2'].font = Font(name='Calibri', size=12)
    
    ws['B3'] = data.get('quality_branch_area', '')
    ws['B3'].font = Font(name='Calibri', size=12)
    
    ws['B4'] = data.get('civil_vendor', '')
    ws['B4'].font = Font(name='Calibri', size=12)
    
    ws['B5'] = data.get('hvac_vendor', '')
    ws['B5'].font = Font(name='Calibri', size=12)
    
    ws['B6'] = data.get('project_start_date', '')
    ws['B6'].font = Font(name='Calibri', size=12)
    
    ws['B7'] = data.get('planned_handover_date', '')
    ws['B7'].font = Font(name='Calibri', size=12)
    
    ws['B8'] = data.get('actual_handover_date', '')
    ws['B8'].font = Font(name='Calibri', size=12)
    
    ws['B9'] = data.get('quality_date_of_visit', '')
    ws['B9'].font = Font(name='Calibri', size=12)
    
    # Start building dynamic sections
    current_row = 10
    
    # ===== SECTION HEADER: Quality Observations =====
    ws[f'A{current_row}'] = 'Sr. No.'
    ws[f'A{current_row}'].font = Font(name='Times New Roman', size=12, bold=True)
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{current_row}'].fill = green_fill
    ws[f'A{current_row}'].border = thin_border
    
    ws.merge_cells(f'B{current_row}:D{current_row}')
    ws[f'B{current_row}'] = 'Quality Observations'
    ws[f'B{current_row}'].font = Font(name='Times New Roman', size=12, bold=True)
    ws[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'B{current_row}'].fill = green_fill
    ws[f'B{current_row}'].border = thin_border
    ws[f'C{current_row}'].border = thin_border
    ws[f'D{current_row}'].border = thin_border
    
    current_row += 1
    
    # Fill Quality Observations data
    filled_count = 0
    for i in range(6):
        obs_value = data.get(f'quality_observation_{i}', '').strip()
        if obs_value:
            ws[f'A{current_row}'] = filled_count + 1
            ws[f'A{current_row}'].font = Font(name='Calibri', size=12)
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'A{current_row}'].border = thin_border
            
            ws[f'B{current_row}'] = obs_value
            ws[f'B{current_row}'].font = Font(name='Calibri', size=12)
            ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws.row_dimensions[current_row].height = None  # Auto height
            
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{current_row}'].border = thin_border
            
            filled_count += 1
            current_row += 1
    
    if filled_count == 0:
        ws[f'A{current_row}'] = 1
        ws[f'A{current_row}'].font = Font(name='Calibri', size=12)
        ws[f'B{current_row}'] = ''
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{current_row}'].border = thin_border
        current_row += 1
    
    # ===== SECTION HEADER: Site Delay Reasons =====
    ws[f'A{current_row}'] = 'Sr. No.'
    ws[f'A{current_row}'].font = Font(name='Times New Roman', size=12, bold=True)
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{current_row}'].fill = green_fill
    ws[f'A{current_row}'].border = thin_border
    
    ws.merge_cells(f'B{current_row}:D{current_row}')
    ws[f'B{current_row}'] = 'Site Delay Reasons'
    ws[f'B{current_row}'].font = Font(name='Times New Roman', size=12, bold=True)
    ws[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'B{current_row}'].fill = green_fill
    ws[f'B{current_row}'].border = thin_border
    ws[f'C{current_row}'].border = thin_border
    ws[f'D{current_row}'].border = thin_border
    
    current_row += 1
    
    # Fill Site Delay Reasons data
    filled_count = 0
    for i in range(6):
        delay_value = data.get(f'site_delay_reason_{i}', '').strip()
        if delay_value:
            ws[f'A{current_row}'] = filled_count + 1
            ws[f'A{current_row}'].font = Font(name='Calibri', size=12)
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'A{current_row}'].border = thin_border
            
            ws[f'B{current_row}'] = delay_value
            ws[f'B{current_row}'].font = Font(name='Calibri', size=12)
            ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            row_height = calculate_row_height(delay_value, font_size=12, cell_width=50)
            ws.row_dimensions[current_row].height = row_height
            
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{current_row}'].border = thin_border
            
            filled_count += 1
            current_row += 1
    
    if filled_count == 0:
        ws[f'A{current_row}'] = 1
        ws[f'A{current_row}'].font = Font(name='Calibri', size=12)
        ws[f'B{current_row}'] = ''
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{current_row}'].border = thin_border
        current_row += 1
    
    # ===== SECTION HEADER: Collaborative Challenges/Inputs =====
    ws[f'A{current_row}'] = 'Sr. No.'
    ws[f'A{current_row}'].font = Font(name='Times New Roman', size=12, bold=True)
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{current_row}'].fill = green_fill
    ws[f'A{current_row}'].border = thin_border
    
    ws.merge_cells(f'B{current_row}:D{current_row}')
    ws[f'B{current_row}'] = 'Collaborative Challenges/Inputs'
    ws[f'B{current_row}'].font = Font(name='Times New Roman', size=12, bold=True)
    ws[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'B{current_row}'].fill = green_fill
    ws[f'B{current_row}'].border = thin_border
    ws[f'C{current_row}'].border = thin_border
    ws[f'D{current_row}'].border = thin_border
    
    current_row += 1
    
    # Fill Collaborative Challenges data
    filled_count = 0
    for i in range(6):
        challenge_value = data.get(f'collaborative_challenge_{i}', '').strip()
        if challenge_value:
            ws[f'A{current_row}'] = filled_count + 1
            ws[f'A{current_row}'].font = Font(name='Calibri', size=12)
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'A{current_row}'].border = thin_border
            
            ws[f'B{current_row}'] = challenge_value
            ws[f'B{current_row}'].font = Font(name='Calibri', size=12)
            ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            row_height = calculate_row_height(challenge_value, font_size=12, cell_width=50)
            ws.row_dimensions[current_row].height = row_height
            
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{current_row}'].border = thin_border
            
            filled_count += 1
            current_row += 1
    
    if filled_count == 0:
        ws[f'A{current_row}'] = 1
        ws[f'A{current_row}'].font = Font(name='Calibri', size=12)
        ws[f'B{current_row}'] = ''
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{current_row}'].border = thin_border
        current_row += 1
    
    # ===== SECTION HEADER: Criticalities =====
    ws[f'A{current_row}'] = 'Sr. No.'
    ws[f'A{current_row}'].font = Font(name='Times New Roman', size=12, bold=True)
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{current_row}'].fill = green_fill
    ws[f'A{current_row}'].border = thin_border
    
    ws.merge_cells(f'B{current_row}:D{current_row}')
    ws[f'B{current_row}'] = 'Criticalities'
    ws[f'B{current_row}'].font = Font(name='Times New Roman', size=12, bold=True)
    ws[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'B{current_row}'].fill = green_fill
    ws[f'B{current_row}'].border = thin_border
    ws[f'C{current_row}'].border = thin_border
    ws[f'D{current_row}'].border = thin_border
    
    current_row += 1
    
    # Fill Criticalities data
    filled_count = 0
    for i in range(6):
        critical_value = data.get(f'criticality_{i}', '').strip()
        if critical_value:
            ws[f'A{current_row}'] = filled_count + 1
            ws[f'A{current_row}'].font = Font(name='Calibri', size=12)
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'A{current_row}'].border = thin_border
            
            ws[f'B{current_row}'] = critical_value
            ws[f'B{current_row}'].font = Font(name='Calibri', size=12)
            ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            row_height = calculate_row_height(critical_value, font_size=12, cell_width=50)
            ws.row_dimensions[current_row].height = row_height
            
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{current_row}'].border = thin_border
            
            filled_count += 1
            current_row += 1
    
    if filled_count == 0:
        ws[f'A{current_row}'] = 1
        ws[f'A{current_row}'].font = Font(name='Calibri', size=12)
        ws[f'B{current_row}'] = ''
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{current_row}'].border = thin_border
        current_row += 1
    
    # ===== SECTION HEADER: Hindrances =====
    ws[f'A{current_row}'] = 'Sr. No.'
    ws[f'A{current_row}'].font = Font(name='Times New Roman', size=12, bold=True)
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{current_row}'].fill = green_fill
    ws[f'A{current_row}'].border = thin_border
    
    ws.merge_cells(f'B{current_row}:D{current_row}')
    ws[f'B{current_row}'] = 'Hindrances'
    ws[f'B{current_row}'].font = Font(name='Times New Roman', size=12, bold=True)
    ws[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'B{current_row}'].fill = green_fill
    ws[f'B{current_row}'].border = thin_border
    ws[f'C{current_row}'].border = thin_border
    ws[f'D{current_row}'].border = thin_border
    
    current_row += 1
    
    # Fill Hindrances data
    filled_count = 0
    for i in range(6):
        hindrance_value = data.get(f'hindrance_{i}', '').strip()
        if hindrance_value:
            ws[f'A{current_row}'] = filled_count + 1
            ws[f'A{current_row}'].font = Font(name='Calibri', size=12)
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'A{current_row}'].border = thin_border
            
            ws[f'B{current_row}'] = hindrance_value
            ws[f'B{current_row}'].font = Font(name='Calibri', size=12)
            ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            row_height = calculate_row_height(hindrance_value, font_size=12, cell_width=50)
            ws.row_dimensions[current_row].height = row_height
            
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{current_row}'].border = thin_border
            
            filled_count += 1
            current_row += 1
    
    if filled_count == 0:
        ws[f'A{current_row}'] = 1
        ws[f'A{current_row}'].font = Font(name='Calibri', size=12)
        ws[f'B{current_row}'] = ''
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{current_row}'].border = thin_border
        current_row += 1
    
    # ===== SECTION HEADER: Others =====
    ws[f'A{current_row}'] = 'Sr. No.'
    ws[f'A{current_row}'].font = Font(name='Times New Roman', size=12, bold=True)
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'A{current_row}'].fill = green_fill
    ws[f'A{current_row}'].border = thin_border
    
    ws.merge_cells(f'B{current_row}:D{current_row}')
    ws[f'B{current_row}'] = 'Others'
    ws[f'B{current_row}'].font = Font(name='Times New Roman', size=12, bold=True)
    ws[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
    ws[f'B{current_row}'].fill = green_fill
    ws[f'B{current_row}'].border = thin_border
    ws[f'C{current_row}'].border = thin_border
    ws[f'D{current_row}'].border = thin_border
    
    current_row += 1
    
    # Fill Others data
    filled_count = 0
    for i in range(6):
        other_value = data.get(f'other_{i}', '').strip()
        if other_value:
            ws[f'A{current_row}'] = filled_count + 1
            ws[f'A{current_row}'].font = Font(name='Calibri', size=12)
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            ws[f'A{current_row}'].border = thin_border
            
            ws[f'B{current_row}'] = other_value
            ws[f'B{current_row}'].font = Font(name='Calibri', size=12)
            ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            row_height = calculate_row_height(other_value, font_size=12, cell_width=50)
            ws.row_dimensions[current_row].height = row_height
            
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{current_row}'].border = thin_border
            
            filled_count += 1
            current_row += 1
    
    if filled_count == 0:
        ws[f'A{current_row}'] = 1
        ws[f'A{current_row}'].font = Font(name='Calibri', size=12)
        ws[f'B{current_row}'] = ''
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{current_row}'].border = thin_border
        current_row += 1
def update_visit_percentages(sol_id, current_visit, data):
    """Update percentages across all visits for the same SOL ID"""
    # This function copies percentages from current visit to previous visits
    # and retrieves percentages from previous visits to add to current visit
    
    for visit in range(1, 5):
        if visit != int(current_visit):
            try:
                blob_path = f"ICICI_Site_Progress_Report/{sol_id}/Visit_{visit}/report.xlsx"
                blob = bucket.blob(blob_path)
                
                if blob.exists():
                    # Download and update the existing visit file
                    excel_buffer = io.BytesIO()
                    blob.download_to_file(excel_buffer)
                    excel_buffer.seek(0)
                    
                    wb = openpyxl.load_workbook(excel_buffer)
                    # Update with current visit percentages
                    # Implementation details...
                    
                    # Save back
                    output_buffer = io.BytesIO()
                    wb.save(output_buffer)
                    output_buffer.seek(0)
                    blob.upload_from_file(output_buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            except Exception as e:
                print(f"Error updating visit {visit}: {e}")
                pass

@app.route('/generate-report', methods=['POST'])
def generate_report():
    try:
        form_data = request.form.to_dict()
        files = request.files
        
        sol_id = form_data.get('sol_id')
        visit_no = form_data.get('visit_no')
        project_name = form_data.get('project_name')
        
        if not sol_id or not visit_no or not project_name:
            return jsonify({'error': 'Missing required fields'}), 400

        work_count = int(form_data.get('work_count', 0))
        quality_count = int(form_data.get('quality_count', 0))
        make_count = int(form_data.get('make_count', 0))
        
        wb = openpyxl.load_workbook(TEMPLATE_PATH)
        
        fill_progress_report(wb, form_data)
        fill_photographs_sheet(wb, form_data, files, sol_id, visit_no)
        fill_quality_sheet(wb, form_data)
        
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        wb.close()
        
        # Save to temporary storage (not final location)
        temp_blob_path = f"ICICI_Site_Progress_Report/temp_drafts/{sol_id}/Visit_{visit_no}/draft.xlsx"
        temp_blob = bucket.blob(temp_blob_path)
        temp_blob.upload_from_file(excel_buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # Save JSON temporarily
        json_blob_path = f"ICICI_Site_Progress_Report/temp_drafts/{sol_id}/Visit_{visit_no}/data.json"
        bucket.blob(json_blob_path).upload_from_string(json.dumps(form_data), content_type='application/json')
        
        gc.collect()
        
        excel_buffer.seek(0)
        return send_file(
            excel_buffer,
            as_attachment=True,
            download_name=f"{project_name}_Visit_{visit_no}_Draft.xlsx",
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': f'Server error: {str(e)}'}), 500

@app.route('/get-reviewers', methods=['POST'])
def get_reviewers():
    data = request.json
    region = data.get('region')
    if region in REGION_REVIEWERS:
        return jsonify({'reviewers': REGION_REVIEWERS[region]})
    return jsonify({'error': 'Invalid region'}), 400

@app.route('/submit-for-review', methods=['POST'])
def submit_for_review():
    try:
        file = request.files['excel_file']
        sol_id = request.form.get('sol_id')
        visit_no = request.form.get('visit_no')
        region = request.form.get('region')
        reviewer_email = request.form.get('reviewer_email')
        user_email = request.form.get('user_email')
        user_name = request.form.get('user_name')
        custom_filename = request.form.get('custom_filename', '').strip()
        user_comments = request.form.get('user_comments', '').strip()  # ✅ NEW
        
        if not all([sol_id, visit_no, region, reviewer_email, user_email]):
            return jsonify({'error': 'Missing required data'}), 400
        
        import uuid
        session_id = str(uuid.uuid4())[:8]
        
        excel_buffer = io.BytesIO()
        file.save(excel_buffer)
        excel_buffer.seek(0)
        
        # ✅ Handle custom filename
        if custom_filename:
            if not custom_filename.endswith('.xlsx'):
                filename = f"{custom_filename}.xlsx"
            else:
                filename = custom_filename
        else:
            filename = f"{sol_id}_Visit_{visit_no}.xlsx"
        
        pending_path = f"ICICI_Site_Progress_Report/{region}/pending_reviews/{session_id}/excel/{filename}"
        pending_blob = bucket.blob(pending_path)
        pending_blob.upload_from_file(excel_buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # ✅ Save metadata with user comments
        metadata = {
            'sol_id': sol_id,
            'visit_no': visit_no,
            'region': region,
            'reviewer_email': reviewer_email,
            'user_email': user_email,
            'user_name': user_name,
            'filename': filename,
            'session_id': session_id,
            'status': 'pending',
            'submitted_at': datetime.now(timezone.utc).isoformat(),
            'user_comments': user_comments  # ✅ NEW: Save comments
        }
        
        metadata_path = f"ICICI_Site_Progress_Report/{region}/pending_reviews/{session_id}/metadata.json"
        bucket.blob(metadata_path).upload_from_string(json.dumps(metadata), content_type='application/json')
        
        return jsonify({'success': True, 'session_id': session_id})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/get-pending-reviews', methods=['POST'])
def get_pending_reviews():
    try:
        data = request.json
        user_email = data.get('user_email')
        region = data.get('region')
        user_type = data.get('user_type')  # 'reviewer' or 'user'
        
        pending_reviews = []
        
        prefix = f"ICICI_Site_Progress_Report/{region}/pending_reviews/"
        blobs = bucket.list_blobs(prefix=prefix)
        
        session_ids = set()
        for blob in blobs:
            if 'metadata.json' in blob.name:
                parts = blob.name.split('/')
                if len(parts) >= 4:
                    session_ids.add(parts[3])
        
        for session_id in session_ids:
            metadata_path = f"ICICI_Site_Progress_Report/{region}/pending_reviews/{session_id}/metadata.json"
            try:
                metadata_blob = bucket.blob(metadata_path)
                if metadata_blob.exists():
                    metadata = json.loads(metadata_blob.download_as_string())
                    
                    if user_type == 'reviewer' and metadata.get('reviewer_email') == user_email:
                        pending_reviews.append(metadata)
                    elif user_type == 'user' and metadata.get('user_email') == user_email:
                        pending_reviews.append(metadata)
            except:
                continue
        
        return jsonify({'reviews': pending_reviews})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/get-completed-reviews', methods=['POST'])
def get_completed_reviews():
    try:
        data = request.json
        user_email = data.get('user_email')
        region = data.get('region')
        user_type = data.get('user_type')
        
        completed_reviews = []
        
        prefix = f"ICICI_Site_Progress_Report/{region}/completed_reviews/"
        blobs = bucket.list_blobs(prefix=prefix)
        
        session_ids = set()
        for blob in blobs:
            if 'metadata.json' in blob.name:
                parts = blob.name.split('/')
                if len(parts) >= 4:
                    session_ids.add(parts[3])
        
        for session_id in session_ids:
            metadata_path = f"ICICI_Site_Progress_Report/{region}/completed_reviews/{session_id}/metadata.json"
            try:
                metadata_blob = bucket.blob(metadata_path)
                if metadata_blob.exists():
                    metadata = json.loads(metadata_blob.download_as_string())
                    
                    if user_type == 'reviewer' and metadata.get('reviewer_email') == user_email:
                        completed_reviews.append(metadata)
                    elif user_type == 'user' and metadata.get('user_email') == user_email:
                        completed_reviews.append(metadata)
            except:
                continue
        
        return jsonify({'reviews': completed_reviews})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/submit-review-decision', methods=['POST'])
def submit_review_decision():
    try:
        data = request.json
        session_id = data.get('session_id')
        region = data.get('region')
        decision = data.get('decision')  # 'approved' or 'rejected'
        comments = data.get('comments', '')
        reviewer_email = data.get('reviewer_email')
        corrected_file_data = data.get('corrected_file')  # base64 if uploaded
        
        # Get metadata
        metadata_path = f"ICICI_Site_Progress_Report/{region}/pending_reviews/{session_id}/metadata.json"
        metadata_blob = bucket.blob(metadata_path)
        metadata = json.loads(metadata_blob.download_as_string())
        
        # Generate review code only if approved
        review_code = None
        if decision == 'approved':
            import random
            import string
            review_code = 'S' + ''.join(random.choices(string.digits + string.ascii_uppercase, k=4))
        
        # Update metadata
        metadata['status'] = decision
        metadata['review_code'] = review_code
        metadata['reviewer_comments'] = comments
        metadata['reviewed_at'] = datetime.now(timezone.utc).isoformat()
        
        # Move to completed
        completed_base = f"ICICI_Site_Progress_Report/{region}/completed_reviews/{session_id}/"
        
        # Copy original Excel
        original_excel_path = f"ICICI_Site_Progress_Report/{region}/pending_reviews/{session_id}/excel/{metadata['filename']}"
        new_excel_path = f"{completed_base}excel/{metadata['filename']}"
        
        source_blob = bucket.blob(original_excel_path)
        bucket.copy_blob(source_blob, bucket, new_excel_path)
        
        # Save corrected file if provided
        if corrected_file_data:
            import base64
            file_bytes = base64.b64decode(corrected_file_data.split(',')[1])
            corrected_path = f"{completed_base}corrected_excel/Corrected_{metadata['filename']}"
            bucket.blob(corrected_path).upload_from_string(file_bytes, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            metadata['corrected_file_path'] = corrected_path
        
        # Save updated metadata
        completed_metadata_path = f"{completed_base}metadata.json"
        bucket.blob(completed_metadata_path).upload_from_string(json.dumps(metadata), content_type='application/json')
        
        # If approved, save to final location and increment count
        if decision == 'approved':
            sol_id = metadata['sol_id']
            visit_no = metadata['visit_no']
            
            # Save to final storage
            final_path = f"ICICI_Site_Progress_Report/{region}/{sol_id}/Visit_{visit_no}/report.xlsx"
            bucket.copy_blob(source_blob, bucket, final_path)
            
            # Make public
            final_blob = bucket.blob(final_path)
            final_blob.make_public()
            
            # Increment user count
            user_email = metadata['user_email']
            count_path = f"ICICI_Site_Progress_Report/{region}/user_stats/{user_email.replace('@', '_at_')}/count.json"
            count_blob = bucket.blob(count_path)
            
            current_count = 0
            if count_blob.exists():
                count_data = json.loads(count_blob.download_as_string())
                current_count = count_data.get('approved_count', 0)
            
            current_count += 1
            count_blob.upload_from_string(json.dumps({'approved_count': current_count, 'last_updated': datetime.now(timezone.utc).isoformat()}), content_type='application/json')
        
        # Delete from pending
        pending_blobs = bucket.list_blobs(prefix=f"ICICI_Site_Progress_Report/{region}/pending_reviews/{session_id}/")
        for blob in pending_blobs:
            blob.delete()
        
        return jsonify({'success': True, 'review_code': review_code})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/get-user-count', methods=['POST'])
def get_user_count():
    try:
        data = request.json
        user_email = data.get('user_email')
        region = data.get('region')
        
        count_path = f"ICICI_Site_Progress_Report/{region}/user_stats/{user_email.replace('@', '_at_')}/count.json"
        count_blob = bucket.blob(count_path)
        
        if count_blob.exists():
            count_data = json.loads(count_blob.download_as_string())
            return jsonify({'count': count_data.get('approved_count', 0)})
        
        return jsonify({'count': 0})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/get-all-approved-reports', methods=['POST'])
def get_all_approved_reports():
    try:
        data = request.json
        region = data.get('region')
        
        reports = []
        
        # List all approved reports in region
        prefix = f"ICICI_Site_Progress_Report/{region}/"
        blobs = bucket.list_blobs(prefix=prefix)
        
        for blob in blobs:
            if '/report.xlsx' in blob.name and '/temp_drafts/' not in blob.name:
                parts = blob.name.split('/')
                if len(parts) >= 4:
                    sol_id = parts[2]
                    visit_folder = parts[3]
                    visit_no = visit_folder.replace('Visit_', '')
                    
                    reports.append({
                        'sol_id': sol_id,
                        'visit_no': visit_no,
                        'download_url': blob.public_url,
                        'filename': f"{sol_id}_Visit_{visit_no}.xlsx"
                    })
        
        return jsonify({'reports': reports})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/download-review-file', methods=['POST'])
def download_review_file():
    try:
        data = request.json
        session_id = data.get('session_id')
        region = data.get('region')
        file_type = data.get('file_type')  # 'original' or 'corrected'
        status = data.get('status')  # 'pending' or 'completed'
        
        if status == 'pending':
            metadata_path = f"ICICI_Site_Progress_Report/{region}/pending_reviews/{session_id}/metadata.json"
        else:
            metadata_path = f"ICICI_Site_Progress_Report/{region}/completed_reviews/{session_id}/metadata.json"
        
        metadata_blob = bucket.blob(metadata_path)
        metadata = json.loads(metadata_blob.download_as_string())
        
        if file_type == 'original':
            if status == 'pending':
                file_path = f"ICICI_Site_Progress_Report/{region}/pending_reviews/{session_id}/excel/{metadata['filename']}"
            else:
                file_path = f"ICICI_Site_Progress_Report/{region}/completed_reviews/{session_id}/excel/{metadata['filename']}"
        else:
            file_path = metadata.get('corrected_file_path')
            if not file_path:
                return jsonify({'error': 'No corrected file available'}), 404
        
        file_blob = bucket.blob(file_path)
        if not file_blob.exists():
            return jsonify({'error': 'File not found'}), 404
        
        file_bytes = file_blob.download_as_bytes()
        file_buffer = io.BytesIO(file_bytes)
        
        return send_file(
            file_buffer,
            as_attachment=True,
            download_name=metadata['filename'],
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/get-session-data', methods=['POST'])
def get_session_data():
    """Fetch user session data from Firebase"""
    try:
        from firebase_admin import firestore
        db_firestore = firestore.client()
        
        data = request.json
        session_id = data.get('session_id')
        
        if not session_id:
            return jsonify({'error': 'No session ID provided'}), 400
        
        # Get session from Firestore
        session_ref = db_firestore.collection('userSessions').document(session_id)
        session_doc = session_ref.get()
        
        if not session_doc.exists:
            return jsonify({'error': 'Session not found'}), 404
        
        session_data = session_doc.to_dict()
        
        # Check if session is expired
        from datetime import datetime
        expires_at = session_data.get('expiresAt')
        
        if expires_at:
            # Convert Firestore timestamp to datetime if needed
            if hasattr(expires_at, 'timestamp'):
                expires_at = datetime.fromtimestamp(expires_at.timestamp())
            
            if expires_at < datetime.now():
                return jsonify({'error': 'Session expired'}), 401
        
        # Update last active time
        session_ref.update({
            'lastActive': firestore.SERVER_TIMESTAMP
        })
        
        return jsonify({
            'email': session_data.get('email'),
            'name': session_data.get('name'),
            'success': True
        })
        
    except Exception as e:
        print(f"Error fetching session: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/verify-user-login', methods=['POST'])
def verify_user_login():
    data = request.json
    email = data.get('email', '').strip()
    password = data.get('password', '').strip()
    
    if not email or not password:
        return jsonify({'success': False, 'error': 'Email and password required'}), 400
    
    try:
        # Check in users collection
        users_ref = db.collection('users')
        query = users_ref.where('email', '==', email).limit(1)
        docs = query.stream()
        
        user_doc = None
        for doc in docs:
            user_doc = doc
            break
        
        if not user_doc:
            return jsonify({'success': False, 'error': 'User not found'}), 401
        
        user_data = user_doc.to_dict()
        stored_password = user_data.get('password', '')

        # ✅ FIX: Verify Password using Bcrypt
        password_valid = False
        
        # Case 1: Password is hashed (starts with $2b$)
        if stored_password.startswith('$2b$'):
            try:
                # Compare the plain text input with the stored hash
                if bcrypt.checkpw(password.encode('utf-8'), stored_password.encode('utf-8')):
                    password_valid = True
            except Exception as e:
                print(f"Hashing error: {e}")
                password_valid = False
        # Case 2: Fallback for plain text passwords (if any exist)
        else:
            if stored_password == password:
                password_valid = True

        if password_valid:
            # Determine region (check if reviewer)
            region = None
            if user_data.get('approved'):
                # User is a reviewer, get their region
                for reg, reviewers in REGION_REVIEWERS.items():
                    if any(r['email'] == email for r in reviewers):
                        region = reg
                        break
            
            return jsonify({
                'success': True,
                'email': email,
                'name': user_data.get('name', ''),
                'region': region
            })
        else:
            return jsonify({'success': False, 'error': 'Incorrect password'}), 401
            
    except Exception as e:
        print(f"Login Error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

def format_date_ddmmyyyy(date_obj):
    """Format datetime to DD/MM/YYYY HH:MM:SS AM/PM format (IST)"""
    if isinstance(date_obj, str):
        try:
            date_obj = datetime.fromisoformat(date_obj.replace('Z', '+00:00'))
        except:
            return date_obj
    
    if isinstance(date_obj, datetime):
        # Convert to IST (UTC+5:30)
        ist = timezone(timedelta(hours=5, minutes=30))
        
        if date_obj.tzinfo is None:
            # Assume UTC if no timezone
            date_obj = date_obj.replace(tzinfo=timezone.utc)
        
        date_obj = date_obj.astimezone(ist)
        
        return date_obj.strftime('%d/%m/%Y, %I:%M:%S %p')
    
    return str(date_obj)

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(debug=False, host='0.0.0.0', port=port)
